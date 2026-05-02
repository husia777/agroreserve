"""
Сервис экспорта данных в Excel/CSV (UC-83).
Выгрузка продуктов, остатков, цен.
"""

import io
from datetime import UTC, datetime
from datetime import date as DateType
from typing import Any, Optional

import structlog
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.models.product import Category, Product

logger = structlog.get_logger(__name__)


# --- Маппинг единиц измерения на русский ---
UNIT_LABELS = {
    "kg": "кг",
    "pcs": "шт",
    "l": "л",
    "box": "ящик",
    "bag": "мешок",
}


async def export_products_excel(
    include_purchase_price: bool = True,
    only_active: bool = False,
    category_id: Optional[str] = None,
) -> io.BytesIO:
    """
    Экспорт товаров в Excel (.xlsx).

    Колонки:
    - Название, Категория, Ед. изм., Остаток, Мин. остаток,
      Цена опт, Цена закупки (опционально), Страна, Условия хранения, Статус
    """
    # Получаем товары
    query: dict[str, Any] = {}
    if only_active:
        query["is_active"] = True
    if category_id:
        query["category_id"] = category_id

    products = await Product.find(query).sort("name").to_list()

    # Получаем категории для маппинга
    categories = await Category.find_all().to_list()
    cat_map = {str(c.id): c.name for c in categories}

    # Создаём книгу Excel
    wb = Workbook()
    ws: _WorksheetOrChartsheetLike = wb.active
    ws.title = "Товары"

    # Стили
    header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="16A34A", end_color="16A34A", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_font = Font(name="Arial", size=10)
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    # Заголовок документа
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = f"Агрорезерв — Каталог товаров ({datetime.now(UTC).strftime('%d.%m.%Y')})"
    title_cell.font = Font(name="Arial", bold=True, size=14, color="16A34A")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    # Заголовки колонок
    headers = [
        ("Название", 35),
        ("Категория", 20),
        ("Ед. изм.", 10),
        ("Остаток", 12),
        ("Мин. остаток", 14),
        ("Цена опт, ₽", 14),
    ]
    if include_purchase_price:
        headers.append(("Цена закупки, ₽", 16))
    headers.extend(
        [
            ("Страна", 15),
            ("Условия хранения", 25),
            ("Статус", 12),
        ]
    )

    for col_idx, (header_text, width) in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=header_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[3].height = 28

    # Данные
    for row_idx, product in enumerate(products, 4):
        cat_name = cat_map.get(str(product.category_id), "—")
        unit_label = UNIT_LABELS.get(product.unit, product.unit)

        row_data = [
            product.name,
            cat_name,
            unit_label,
            product.stock_qty,
            product.min_stock_qty,
            product.price_wholesale,
        ]
        if include_purchase_price:
            row_data.append(product.cost_price or 0)
        row_data.extend(
            [
                product.origin_country or "—",
                product.storage_conditions or "—",
                "Активен" if product.is_active else "Неактивен",
            ]
        )

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = cell_font
            cell.border = thin_border
            if isinstance(value, (int, float)):
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = "#,##0.00" if isinstance(value, float) else "#,##0"

        # Подсветка критических остатков
        stock_qty = product.stock_qty or 0
        min_stock = product.min_stock_qty or 0
        if stock_qty <= 0:
            ws.cell(row=row_idx, column=4).fill = PatternFill(
                start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"
            )
        elif stock_qty <= min_stock:
            ws.cell(row=row_idx, column=4).fill = PatternFill(
                start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"
            )

    # Итого строка
    total_row = len(products) + 4
    ws.cell(row=total_row, column=1, value=f"Всего товаров: {len(products)}")
    ws.cell(row=total_row, column=1).font = Font(name="Arial", bold=True, size=10)

    # Автофильтр
    last_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A3:{last_col}{total_row - 1}"

    # Закрепляем заголовок
    ws.freeze_panes = "A4"

    # Сохраняем в буфер
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    logger.info("Экспорт товаров в Excel", count=len(products), include_purchase=include_purchase_price)
    return buffer


async def export_products_csv(
    include_purchase_price: bool = True,
    only_active: bool = False,
) -> io.StringIO:
    """
    Экспорт товаров в CSV (UTF-8 с BOM для корректного открытия в Excel).
    """
    import csv

    products = await Product.find({"is_active": True} if only_active else {}).sort("name").to_list()

    categories = await Category.find_all().to_list()
    cat_map = {str(c.id): c.name for c in categories}

    buffer = io.StringIO()
    # BOM для Excel
    buffer.write("\ufeff")

    headers = ["Название", "Категория", "Ед. изм.", "Остаток", "Мин. остаток", "Цена опт"]
    if include_purchase_price:
        headers.append("Цена закупки")
    headers.extend(["Страна", "Статус"])

    writer = csv.writer(buffer, delimiter=";")
    writer.writerow(headers)

    for product in products:
        row = [
            product.name,
            cat_map.get(str(product.category_id), "—"),
            UNIT_LABELS.get(product.unit, product.unit),
            product.stock_qty,
            product.min_stock_qty,
            product.price_wholesale,
        ]
        if include_purchase_price:
            row.append(product.cost_price or 0)
        row.extend(
            [
                product.origin_country or "—",
                "Активен" if product.is_active else "Неактивен",
            ]
        )
        writer.writerow(row)

    buffer.seek(0)
    logger.info("Экспорт товаров в CSV", count=len(products))
    return buffer


# ───────────────────────────────────────────────────────────────
# UC-83: Экспорт заказов за период в Excel (с маржой)
# ───────────────────────────────────────────────────────────────

# Маппинг статусов заказа на русский
ORDER_STATUS_LABELS = {
    "new": "Новый",
    "confirmed": "Подтверждён",
    "processing": "В обработке",
    "ready": "Готов к отгрузке",
    "in_delivery": "В доставке",
    "delivered": "Доставлен",
    "completed": "Завершён",
    "cancelled": "Отменён",
    "returned": "Возврат",
}

PAYMENT_STATUS_LABELS = {
    "unpaid": "Не оплачен",
    "partial": "Частично",
    "paid": "Оплачен",
    "overdue": "Просрочен",
    "refunded": "Возвращён",
}

PAYMENT_METHOD_LABELS = {
    "cash": "Наличные",
    "card": "Карта",
    "bank_transfer": "Безнал",
    "credit": "Отсрочка",
}


async def export_orders_excel(
    date_from: "DateType",
    date_to: "DateType",
    statuses: Optional[list[str]] = None,
    client_id: Optional[str] = None,
    only_paid: bool = False,
) -> io.BytesIO:
    """
    UC-83: Экспорт заказов за период в Excel — 3 листа (Заказы, Позиции, Итоги).

    Args:
        date_from: Дата начала периода (включительно)
        date_to: Дата окончания периода (включительно)
        statuses: Фильтр по статусам (если None — все)
        client_id: Фильтр по конкретному клиенту (если None — все)
        only_paid: Только оплаченные заказы

    Returns:
        Буфер с .xlsx файлом
    """
    from datetime import datetime as _dt
    from datetime import time as _time

    from app.models.order import Order

    # Преобразуем даты в datetime для фильтра по created_at
    dt_from = _dt.combine(date_from, _time.min).replace(tzinfo=UTC)
    dt_to = _dt.combine(date_to, _time.max).replace(tzinfo=UTC)

    query: dict[str, Any] = {
        "created_at": {"$gte": dt_from, "$lte": dt_to},
    }
    if statuses:
        query["status"] = {"$in": statuses}
    if client_id:
        try:
            query["client_id.$id"] = PydanticObjectId(client_id)
        except Exception:
            pass
    if only_paid:
        query["payment_status"] = "paid"

    orders = await Order.find(query, fetch_links=False).sort("-created_at").to_list()

    # Создаём книгу
    wb = Workbook()

    # Общие стили
    header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="16A34A", end_color="16A34A", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_font = Font(name="Arial", size=10)
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    # ───────── Лист 1: Заказы ─────────
    ws1 = wb.active
    ws1.title = "Заказы"

    ws1.merge_cells("A1:L1")
    t1 = ws1["A1"]
    t1.value = f"Агрорезерв — Заказы с {date_from.strftime('%d.%m.%Y')} " f"по {date_to.strftime('%d.%m.%Y')}"
    t1.font = Font(name="Arial", bold=True, size=14, color="16A34A")
    ws1.row_dimensions[1].height = 28

    orders_headers = [
        ("№ заказа", 16),
        ("Дата", 12),
        ("Клиент", 30),
        ("Телефон", 15),
        ("Статус", 18),
        ("Способ оплаты", 15),
        ("Оплата", 14),
        ("Сумма, ₽", 14),
        ("Себест., ₽", 14),
        ("Маржа, ₽", 14),
        ("Маржа, %", 11),
        ("Доставка", 12),
    ]
    for col_idx, (text, width) in enumerate(orders_headers, 1):
        c = ws1.cell(row=3, column=col_idx, value=text)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_alignment
        c.border = thin_border
        ws1.column_dimensions[get_column_letter(col_idx)].width = width
    ws1.row_dimensions[3].height = 30

    total_revenue = 0.0
    total_cost = 0.0
    total_discount = 0.0

    for row_idx, order in enumerate(orders, 4):
        # Считаем себестоимость
        cost_sum = sum((item.cost_price or 0) * (item.actual_qty or item.ordered_qty or 0) for item in order.items)
        margin_rub = order.total - cost_sum
        margin_pct = round(margin_rub / order.total * 100, 1) if order.total > 0 else 0.0

        total_revenue += order.total
        total_cost += cost_sum
        total_discount += order.discount or 0

        row = [
            order.order_number,
            order.created_at.date() if order.created_at else None,
            order.client_name,
            order.client_phone,
            ORDER_STATUS_LABELS.get(
                order.status.value if hasattr(order.status, "value") else str(order.status),
                str(order.status),
            ),
            PAYMENT_METHOD_LABELS.get(
                order.payment_method.value if hasattr(order.payment_method, "value") else str(order.payment_method),
                str(order.payment_method),
            ),
            PAYMENT_STATUS_LABELS.get(
                order.payment_status.value if hasattr(order.payment_status, "value") else str(order.payment_status),
                str(order.payment_status),
            ),
            round(order.total, 2),
            round(cost_sum, 2),
            round(margin_rub, 2),
            margin_pct,
            order.delivery_date.strftime("%d.%m.%Y") if order.delivery_date else "—",
        ]

        for col_idx, value in enumerate(row, 1):
            c = ws1.cell(row=row_idx, column=col_idx, value=value)
            c.font = cell_font
            c.border = thin_border
            if isinstance(value, (int, float)) and not isinstance(value, bool):
                c.alignment = Alignment(horizontal="right")
                c.number_format = "#,##0.00" if isinstance(value, float) else "#,##0"

        # Подсветка неоплаченных и просроченных
        ps_str = order.payment_status.value if hasattr(order.payment_status, "value") else str(order.payment_status)
        if ps_str == "overdue":
            ws1.cell(row=row_idx, column=7).fill = PatternFill(
                start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"
            )
        elif ps_str == "unpaid":
            ws1.cell(row=row_idx, column=7).fill = PatternFill(
                start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"
            )

    # Строка ИТОГО в листе Заказы
    last_row = len(orders) + 4
    ws1.cell(row=last_row, column=1, value="ИТОГО").font = Font(name="Arial", bold=True, size=11)
    ws1.cell(row=last_row, column=8, value=round(total_revenue, 2)).font = Font(name="Arial", bold=True)
    ws1.cell(row=last_row, column=8).number_format = "#,##0.00"
    ws1.cell(row=last_row, column=9, value=round(total_cost, 2)).font = Font(name="Arial", bold=True)
    ws1.cell(row=last_row, column=9).number_format = "#,##0.00"
    ws1.cell(row=last_row, column=10, value=round(total_revenue - total_cost, 2)).font = Font(name="Arial", bold=True)
    ws1.cell(row=last_row, column=10).number_format = "#,##0.00"

    ws1.auto_filter.ref = f"A3:L{last_row - 1}" if orders else "A3:L3"
    ws1.freeze_panes = "A4"

    # ───────── Лист 2: Позиции ─────────
    ws2 = wb.create_sheet(title="Позиции")

    ws2.merge_cells("A1:J1")
    t2 = ws2["A1"]
    t2.value = "Детализация по позициям"
    t2.font = Font(name="Arial", bold=True, size=14, color="16A34A")
    ws2.row_dimensions[1].height = 28

    items_headers = [
        ("№ заказа", 16),
        ("Дата", 12),
        ("Клиент", 30),
        ("Товар", 35),
        ("Кол-во", 10),
        ("Ед.", 8),
        ("Цена, ₽", 12),
        ("Себест., ₽", 12),
        ("Сумма, ₽", 14),
        ("Маржа, ₽", 14),
    ]
    for col_idx, (text, width) in enumerate(items_headers, 1):
        c = ws2.cell(row=3, column=col_idx, value=text)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_alignment
        c.border = thin_border
        ws2.column_dimensions[get_column_letter(col_idx)].width = width
    ws2.row_dimensions[3].height = 30

    row_idx = 4
    for order in orders:
        for item in order.items:
            qty = item.actual_qty or item.ordered_qty or 0
            cost = (item.cost_price or 0) * qty
            row = [
                order.order_number,
                order.created_at.date() if order.created_at else None,
                order.client_name,
                item.product_name,
                qty,
                UNIT_LABELS.get(item.unit, item.unit),
                round(item.price, 2),
                round(item.cost_price or 0, 2),
                round(item.total, 2),
                round(item.total - cost, 2),
            ]
            for col_idx, value in enumerate(row, 1):
                c = ws2.cell(row=row_idx, column=col_idx, value=value)
                c.font = cell_font
                c.border = thin_border
                if isinstance(value, (int, float)) and not isinstance(value, bool):
                    c.alignment = Alignment(horizontal="right")
                    c.number_format = "#,##0.00" if isinstance(value, float) else "#,##0"
            row_idx += 1

    ws2.auto_filter.ref = f"A3:J{row_idx - 1}" if row_idx > 4 else "A3:J3"
    ws2.freeze_panes = "A4"

    # ───────── Лист 3: Итоги ─────────
    ws3 = wb.create_sheet(title="Итоги")

    ws3.merge_cells("A1:B1")
    t3 = ws3["A1"]
    t3.value = f"Сводка за {date_from.strftime('%d.%m.%Y')} — {date_to.strftime('%d.%m.%Y')}"
    t3.font = Font(name="Arial", bold=True, size=14, color="16A34A")
    ws3.row_dimensions[1].height = 28
    ws3.column_dimensions["A"].width = 35
    ws3.column_dimensions["B"].width = 20

    total_margin = total_revenue - total_cost
    total_margin_pct = round(total_margin / total_revenue * 100, 1) if total_revenue > 0 else 0.0
    avg_check = round(total_revenue / len(orders), 2) if orders else 0.0

    # Группируем по статусу
    status_counts: dict[str, int] = {}
    for o in orders:
        s = o.status.value if hasattr(o.status, "value") else str(o.status)
        status_counts[s] = status_counts.get(s, 0) + 1

    summary_rows = [
        ("Всего заказов", len(orders)),
        ("Сумма выручки, ₽", round(total_revenue, 2)),
        ("Себестоимость, ₽", round(total_cost, 2)),
        ("Маржа, ₽", round(total_margin, 2)),
        ("Маржа, %", total_margin_pct),
        ("Средний чек, ₽", avg_check),
        ("Сумма скидок, ₽", round(total_discount, 2)),
        ("", ""),
        ("— По статусам —", ""),
    ]
    for status, cnt in sorted(status_counts.items(), key=lambda x: -x[1]):
        summary_rows.append((ORDER_STATUS_LABELS.get(status, status), cnt))

    for r_idx, (label, value) in enumerate(summary_rows, 3):
        c1 = ws3.cell(row=r_idx, column=1, value=label)
        c2 = ws3.cell(row=r_idx, column=2, value=value)
        c1.font = Font(
            name="Arial",
            size=11,
            bold=label.startswith("—") or label in ("Всего заказов", "Сумма выручки, ₽", "Маржа, ₽"),
        )
        c2.font = Font(name="Arial", size=11)
        if isinstance(value, float):
            c2.number_format = "#,##0.00"
            c2.alignment = Alignment(horizontal="right")
        elif isinstance(value, int):
            c2.number_format = "#,##0"
            c2.alignment = Alignment(horizontal="right")

    # Сохраняем
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    logger.info(
        "Экспорт заказов в Excel",
        period=f"{date_from} — {date_to}",
        count=len(orders),
        revenue=total_revenue,
        margin=total_margin,
    )
    return buffer
